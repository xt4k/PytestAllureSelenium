import openpyxl

book = openpyxl.load_workbook("C:\\Users\\Admin\\PycharmProjects\\FirstFramework\\test_data\\data.xlsx")

sheet = book.active

cell = sheet.cell(row=2,column=2)

print("cell read: ", cell.value)

sheet.cell(row=4,column=4).value = "data from python code."

cell2 = sheet.cell(row=4,column=4)

print("cell(4,4) read: ", cell2.value)

print("cell(H5) read: ", sheet['H5'].value)

for i in range(1, sheet.max_row+1):
    print(sheet.cell(row=i,column=8).value)


