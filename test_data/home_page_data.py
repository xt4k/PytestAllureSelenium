import openpyxl


class HomePageData:

    test_HomePage_data = [
            {
                "first_name": "first_1",
                "last_name": "last_1",
                "email": "e@mai.l1",
                "password": "Password_1",
                "gender": "Male"
            },{
                "first_name": "first_2",
                "last_name": "last_2",
                "email": "e@mai.l2",
                "password": "Password_2",
                "gender": "Female"
            }
    ]

    @staticmethod
    def get_test_data(test_case_name):
        print("begin read excel data")
        Dict = {}
        book = openpyxl.load_workbook(r"C:/Users/Admin/PycharmProjects/PytestAllureSelenium/PytestAllureSelenium/test_data/data2.xlsx")
        #book = openpyxl.load_workbook("C:\\Users\\Admin\\PycharmProjects\\PytestAllureSelenium\\test_data\\data2.xlsx")
        print("book_info sheetnames", book.sheetnames)
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i,column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    print("sheet_cell: ",sheet.cell(row=i, column=8).value)
                    Dict[sheet.cell(row=1,column=j).value] = sheet.cell(row=i,column=j).value
        return [Dict]
